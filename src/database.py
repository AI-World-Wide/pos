# -*- coding: utf-8 -*-
"""SQLAlchemy engine + session factory + init_db().

SQLite-only. No PostgreSQL connections — the production Dell has another POS
on its own PostgreSQL install that must not be touched.
"""
from __future__ import annotations

import configparser
import sys
from pathlib import Path

from sqlalchemy import create_engine, event
from sqlalchemy.engine import Engine
from sqlalchemy.orm import sessionmaker, Session

# When frozen (PyInstaller), data/ lives next to the .exe, not inside _MEIPASS.
if getattr(sys, "frozen", False):
    RUNTIME_DIR = Path(sys.executable).resolve().parent
    BUNDLE_DIR = Path(sys._MEIPASS)
else:
    RUNTIME_DIR = Path(__file__).resolve().parent.parent
    BUNDLE_DIR = RUNTIME_DIR

PROJECT_ROOT = RUNTIME_DIR
CONFIG_PATH = BUNDLE_DIR / "config.ini"
if (RUNTIME_DIR / "config.ini").exists():
    CONFIG_PATH = RUNTIME_DIR / "config.ini"


def _load_db_path() -> Path:
    config = configparser.ConfigParser()
    config.read(CONFIG_PATH, encoding="utf-8")
    rel = config.get("database", "path", fallback="data/pos.db")
    return RUNTIME_DIR / rel


DB_PATH = _load_db_path()
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

# SQLite URL with relative path resolved to absolute for cross-platform safety.
engine: Engine = create_engine(
    f"sqlite:///{DB_PATH.as_posix()}",
    future=True,
    echo=False,
    connect_args={"check_same_thread": False},
)


# Enable foreign keys on every SQLite connection (off by default in SQLite).
@event.listens_for(engine, "connect")
def _enable_sqlite_fk(dbapi_conn, connection_record):
    cursor = dbapi_conn.cursor()
    cursor.execute("PRAGMA foreign_keys=ON")
    cursor.execute("PRAGMA journal_mode=WAL")
    cursor.close()


SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)


def get_session() -> Session:
    """Open a new session. Caller is responsible for closing/committing."""
    return SessionLocal()


def init_db() -> None:
    """Create all tables if they don't exist, then seed users + permissions + areas + items."""
    from src.models import Base, Item
    from src.seed_users import seed_users_and_permissions
    from src.seed_tables import seed_areas_and_tables

    Base.metadata.create_all(bind=engine)
    with SessionLocal() as session:
        seed_users_and_permissions(session)
        seed_areas_and_tables(session)
        session.commit()

    # Auto-import items from seed XLSX if the items table is empty
    with SessionLocal() as session:
        if session.query(Item).count() == 0:
            _auto_import_items(session)


def _auto_import_items(session) -> None:
    """Import items from the seed XLSX file if present."""
    import logging
    logger = logging.getLogger(__name__)

    seed_path = RUNTIME_DIR / "data" / "seed" / "Item List_20-05-2026.xlsx"
    if not seed_path.exists():
        logger.info("No seed XLSX found at %s — skipping auto-import", seed_path)
        return

    try:
        from openpyxl import load_workbook
        from src.models import Category
        from src.translations.ar import CATEGORY_AR_NAMES, CATEGORY_IS_BEVERAGE, CATEGORY_SORT_ORDER

        # Ensure categories exist
        existing_cats = {c.name_en: c for c in session.query(Category).all()}
        for name_en, name_ar in CATEGORY_AR_NAMES.items():
            if name_en not in existing_cats:
                cat = Category(
                    name_en=name_en, name_ar=name_ar,
                    sort_order=CATEGORY_SORT_ORDER.get(name_en, 9999),
                    is_beverage=CATEGORY_IS_BEVERAGE.get(name_en, 0),
                    visible=1,
                )
                session.add(cat)
                existing_cats[name_en] = cat
        session.flush()

        wb = load_workbook(filename=str(seed_path), data_only=True, read_only=True)
        ws = wb.active

        from src.models import Item
        count = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row is None or all(v is None for v in row):
                continue
            name_ar = (str(row[1]).strip() if row[1] else "")
            price_raw = row[3]
            category_en = (str(row[5]).strip() if row[5] else "")
            visible_raw = (str(row[8]).strip().upper() if row[8] else "")

            if not name_ar or category_en not in existing_cats:
                continue
            try:
                price = float(price_raw) if price_raw is not None else 0.0
            except (TypeError, ValueError):
                continue

            session.add(Item(
                category_id=existing_cats[category_en].id,
                name_ar=name_ar,
                name_en=(str(row[0]).strip() if row[0] else None),
                price_inclusive=price,
                barcode=(str(row[4]).strip() if row[4] else None),
                kitchen_station="",
                visible=1 if visible_raw == "YES" else 0,
            ))
            count += 1

        session.commit()
        logger.info("Auto-imported %d items from seed XLSX", count)
    except Exception as e:
        logger.error("Auto-import failed: %s", e)
        session.rollback()
