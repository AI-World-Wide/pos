# -*- coding: utf-8 -*-
"""Idempotent XLSX → SQLite seeder.

Reads `data/seed/Item List_20-05-2026.xlsx` and upserts 6 categories + 195
items into `data/pos.db`. Prices are imported as VAT-inclusive sticker values.

Run:
    python scripts/import_items.py
"""
from __future__ import annotations

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook  # noqa: E402

from src.database import SessionLocal, init_db  # noqa: E402
from src.models import Category, Item  # noqa: E402
from src.translations.ar import (  # noqa: E402
    CATEGORY_AR_NAMES,
    CATEGORY_IS_BEVERAGE,
    CATEGORY_SORT_ORDER,
)

SEED_PATH = PROJECT_ROOT / "data" / "seed" / "Item List_20-05-2026.xlsx"

# XLSX has two header rows (row 1 = title, row 2 = column headers). Data starts row 3.
DATA_START_ROW = 3

# Column mapping (1-indexed).
COL_NAME_EN = 1   # A
COL_NAME_AR = 2   # B
COL_PRICE = 4     # D
COL_BARCODE = 5   # E
COL_CATEGORY = 6  # F
COL_PRINTER = 8   # H
COL_VISIBLE = 9   # I

# Printer name "AnyDesk Printer" is a stray; treat as no station so it routes
# to the default receipt printer at order time (Phase 3 behavior).
PRINTER_DEFAULTS = {"": "", "AnyDesk Printer": ""}


def _normalize_printer(raw: str) -> str:
    raw = (raw or "").strip()
    return PRINTER_DEFAULTS.get(raw, raw)


def _upsert_categories(session) -> dict[str, Category]:
    """Make sure every brief-defined category exists. Return name_en → row."""
    existing = {c.name_en: c for c in session.query(Category).all()}
    for name_en, name_ar in CATEGORY_AR_NAMES.items():
        cat = existing.get(name_en)
        if cat is None:
            cat = Category(
                name_en=name_en,
                name_ar=name_ar,
                sort_order=CATEGORY_SORT_ORDER.get(name_en, 9999),
                is_beverage=CATEGORY_IS_BEVERAGE.get(name_en, 0),
                visible=1,
            )
            session.add(cat)
            existing[name_en] = cat
        else:
            cat.name_ar = name_ar
            cat.sort_order = CATEGORY_SORT_ORDER.get(name_en, 9999)
            cat.is_beverage = CATEGORY_IS_BEVERAGE.get(name_en, 0)
    session.flush()  # make sure new categories have IDs before items reference them
    return existing


def _upsert_items(session, categories: dict[str, Category]) -> tuple[int, int]:
    if not SEED_PATH.exists():
        raise FileNotFoundError(f"Seed file not found: {SEED_PATH}")

    wb = load_workbook(filename=str(SEED_PATH), data_only=True, read_only=True)
    ws = wb.active

    added = updated = 0
    skipped: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW):
        if row is None or all(v is None for v in row):
            continue

        name_en = (str(row[COL_NAME_EN - 1]).strip() if row[COL_NAME_EN - 1] else "")
        name_ar = (str(row[COL_NAME_AR - 1]).strip() if row[COL_NAME_AR - 1] else "")
        price_raw = row[COL_PRICE - 1]
        category_en = (str(row[COL_CATEGORY - 1]).strip() if row[COL_CATEGORY - 1] else "")
        printer_raw = (str(row[COL_PRINTER - 1]).strip() if row[COL_PRINTER - 1] else "")
        visible_raw = (str(row[COL_VISIBLE - 1]).strip().upper() if row[COL_VISIBLE - 1] else "")
        barcode = (str(row[COL_BARCODE - 1]).strip() if row[COL_BARCODE - 1] else "")

        if not name_ar:
            skipped.append(f"row {row_idx}: missing Arabic name")
            continue
        if category_en not in categories:
            skipped.append(f"row {row_idx}: unknown category '{category_en}'")
            continue
        try:
            price = float(price_raw) if price_raw is not None else 0.0
        except (TypeError, ValueError):
            skipped.append(f"row {row_idx}: invalid price '{price_raw}'")
            continue

        cat = categories[category_en]
        # Idempotency key: (category, name_en, name_ar, price). The XLSX has 3
        # legit cases where (category, name_en) repeats — fall back to the full
        # quadruple to keep imports idempotent without losing rows.
        existing = (
            session.query(Item)
            .filter(
                Item.category_id == cat.id,
                Item.name_en == (name_en or None),
                Item.name_ar == name_ar,
                Item.price_inclusive == price,
            )
            .first()
        )
        if existing is None:
            session.add(Item(
                category_id=cat.id,
                name_ar=name_ar,
                name_en=name_en or None,
                price_inclusive=price,
                barcode=barcode or None,
                kitchen_station=_normalize_printer(printer_raw),
                visible=1 if visible_raw == "YES" else 0,
            ))
            added += 1
        else:
            existing.name_ar = name_ar
            existing.price_inclusive = price
            existing.barcode = barcode or None
            existing.kitchen_station = _normalize_printer(printer_raw)
            existing.visible = 1 if visible_raw == "YES" else 0
            updated += 1

    if skipped:
        print(f"  Skipped {len(skipped)} row(s):")
        for s in skipped:
            print(f"   - {s}")
    return added, updated


def main() -> None:
    init_db()
    with SessionLocal() as session:
        categories = _upsert_categories(session)
        added, updated = _upsert_items(session, categories)
        session.commit()

    cat_count = len(categories)
    total = added + updated
    print(f"Imported: {total} items ({added} new, {updated} updated) across {cat_count} categories.")


if __name__ == "__main__":
    main()
